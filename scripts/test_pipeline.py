import sys
import json
import pytest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

# ── AdapterBase tests ────────────────────────────────────────────────────────

from pipeline.adapter_base import AdapterBase


class ConcreteAdapter(AdapterBase):
    SOURCE_ID   = "test_src"
    SOURCE_NAME = "Test Source"
    PROGRAM     = "MWBE"
    GEOGRAPHY   = "TEST"
    CONFIDENCE  = "confirmed_black"
    FIELD_MAP   = {
        "BizName":  "business_name",
        "OwnerNm":  "owner_name",
        "ZipCode":  "address_zip",
    }

    def fetch(self):
        return [
            {"BizName": "Acme LLC", "OwnerNm": "Jane Doe",
             "ZipCode": "10001", "ExtraCol": "extra_value"},
        ]

    def parse(self, raw):
        return [self.map_record(row) for row in raw]


def test_adapter_run_returns_list():
    adapter = ConcreteAdapter()
    records = adapter.run()
    assert isinstance(records, list)
    assert len(records) == 1


def test_adapter_map_record_applies_field_map():
    adapter = ConcreteAdapter()
    records = adapter.run()
    assert records[0]["business_name"] == "Acme LLC"
    assert records[0]["owner_name"] == "Jane Doe"
    assert records[0]["address_zip"] == "10001"


def test_adapter_map_record_puts_unmapped_in_source_fields():
    adapter = ConcreteAdapter()
    records = adapter.run()
    sf = records[0]["source_fields"]
    assert sf["ExtraCol"] == "extra_value"


def test_adapter_map_record_fills_missing_bbrt_fields_with_empty_string():
    adapter = ConcreteAdapter()
    records = adapter.run()
    assert records[0]["address_street"] == ""
    assert records[0]["latitude"] == ""


def test_adapter_map_record_sets_data_source():
    adapter = ConcreteAdapter()
    records = adapter.run()
    assert records[0]["data_source"] == "Test Source"


def test_adapter_missing_fetch_raises():
    with pytest.raises(TypeError):
        class BadAdapter(AdapterBase):
            SOURCE_ID = "bad"
            SOURCE_NAME = "Bad"
            PROGRAM = "MWBE"
            GEOGRAPHY = "X"
            CONFIDENCE = "confirmed_black"
            FIELD_MAP = {}
            # missing fetch and parse
        BadAdapter()


# ── db.py tests ──────────────────────────────────────────────────────────────

import duckdb
import json
from pipeline.db import open_db, upsert_source, write_businesses, write_snapshot_meta, get_registry, upsert_registry, carry_forward_records


@pytest.fixture
def tmp_db(tmp_path):
    db_path = tmp_path / "test.duckdb"
    con = open_db(db_path)
    yield con
    con.close()


class _MockAdapter:
    SOURCE_ID   = "src_a"
    SOURCE_NAME = "Source A"
    PROGRAM     = "MWBE"
    GEOGRAPHY   = "NYC"
    CONFIDENCE  = "confirmed_black"


def _make_record(**kwargs):
    base = {
        "business_id": "uuid-1", "business_name": "Biz A",
        "owner_name": "", "year_founded": "2010",
        "address_street": "1 Main St", "address_city": "Brooklyn",
        "address_state": "New York", "address_zip": "11201",
        "latitude": "40.68", "longitude": "-73.94",
        "industry": "Services", "naics_code": "561990",
        "certification": "MBE", "description": "A business.",
        "website": "", "phone": "", "email": "",
        "instagram_handle": "", "facebook_url": "", "tiktok_handle": "",
        "yelp_url": "", "google_maps_url": "",
        "discloses_google_maps": "", "discloses_yelp": "", "discloses_instagram": "",
        "data_source": "Source A", "last_verified": "2025-09-09",
        "source_id": "src_a", "source_business_id": "ACC001",
        "confidence": "confirmed_black", "source_fields": {"ExtraCol": "val"},
        "canonical_name": "biz a", "canonical_zip": "11201",
    }
    base.update(kwargs)
    return base


def test_open_db_creates_all_tables(tmp_db):
    tables = {row[0] for row in tmp_db.execute(
        "SELECT table_name FROM information_schema.tables WHERE table_schema='main'"
    ).fetchall()}
    assert {"sources", "snapshots", "businesses", "business_registry", "field_catalog"} <= tables


def test_upsert_source_inserts_new(tmp_db):
    upsert_source(tmp_db, _MockAdapter())
    count = tmp_db.execute("SELECT COUNT(*) FROM sources WHERE source_id='src_a'").fetchone()[0]
    assert count == 1


def test_upsert_source_is_idempotent(tmp_db):
    upsert_source(tmp_db, _MockAdapter())
    upsert_source(tmp_db, _MockAdapter())
    count = tmp_db.execute("SELECT COUNT(*) FROM sources").fetchone()[0]
    assert count == 1


def test_write_businesses_inserts_records(tmp_db):
    records = [_make_record()]
    write_businesses(tmp_db, records, "2026-Q2")
    count = tmp_db.execute("SELECT COUNT(*) FROM businesses").fetchone()[0]
    assert count == 1


def test_write_businesses_stores_source_fields_as_json(tmp_db):
    records = [_make_record()]
    write_businesses(tmp_db, records, "2026-Q2")
    sf = tmp_db.execute("SELECT source_fields FROM businesses").fetchone()[0]
    assert json.loads(sf)["ExtraCol"] == "val"


def test_write_snapshot_meta_inserts_row(tmp_db):
    write_snapshot_meta(tmp_db, "2026-Q2", 10, 2, ["src_a"], ["src_b"])
    row = tmp_db.execute("SELECT * FROM snapshots WHERE snapshot_id='2026-Q2'").fetchone()
    assert row is not None
    assert row[2] == 10   # records_added
    assert row[3] == 2    # records_dropped


def test_get_registry_returns_empty_initially(tmp_db):
    result = get_registry(tmp_db)
    assert result == []


def test_upsert_registry_inserts_new_entries(tmp_db):
    entries = [{"business_id": "uuid-1", "canonical_name": "biz a",
                "canonical_zip": "11201", "source_id": "src_a",
                "source_business_id": "ACC001"}]
    upsert_registry(tmp_db, "2026-Q2", entries)
    result = get_registry(tmp_db)
    assert len(result) == 1
    assert result[0]["business_id"] == "uuid-1"


def test_upsert_registry_updates_last_seen(tmp_db):
    entries = [{"business_id": "uuid-1", "canonical_name": "biz a",
                "canonical_zip": "11201", "source_id": "src_a",
                "source_business_id": "ACC001"}]
    upsert_registry(tmp_db, "2026-Q1", entries)
    upsert_registry(tmp_db, "2026-Q2", entries)
    result = get_registry(tmp_db)
    assert result[0]["last_seen"] == "2026-Q2"
    assert result[0]["first_seen"] == "2026-Q1"


# ── carry-forward tests ──────────────────────────────────────────────────────

def test_carry_forward_returns_records_for_unrun_sources(tmp_db):
    write_businesses(tmp_db, [
        _make_record(business_id="uuid-1", source_id="src_a", business_name="A Co"),
        _make_record(business_id="uuid-2", source_id="src_b", business_name="B Co"),
    ], "2026-Q2")
    write_snapshot_meta(tmp_db, "2026-Q2", 2, 0, ["src_a", "src_b"], [])

    # New run: only src_a succeeded; src_b did not run this cycle.
    carried = carry_forward_records(tmp_db, "2026-Q3", {"src_a"})

    ids = {r["business_id"] for r in carried}
    assert ids == {"uuid-2"}                       # only the un-run source carried
    assert carried[0]["source_id"] == "src_b"
    assert carried[0]["business_name"] == "B Co"   # fields preserved verbatim


def test_snapshot_meta_records_carried_forward_sources(tmp_db):
    write_snapshot_meta(tmp_db, "2026-Q3", 10, 0, ["src_a"], [],
                        sources_carried_forward=["src_b", "src_c"])
    row = tmp_db.execute(
        "SELECT sources_carried_forward FROM snapshots WHERE snapshot_id='2026-Q3'"
    ).fetchone()
    assert json.loads(row[0]) == ["src_b", "src_c"]


def test_snapshot_meta_carried_forward_defaults_empty(tmp_db):
    write_snapshot_meta(tmp_db, "2026-Q3", 10, 0, ["src_a"], [])
    row = tmp_db.execute(
        "SELECT sources_carried_forward FROM snapshots WHERE snapshot_id='2026-Q3'"
    ).fetchone()
    assert json.loads(row[0]) == []


def test_carry_forward_empty_when_all_sources_ran(tmp_db):
    write_businesses(tmp_db, [
        _make_record(business_id="uuid-1", source_id="src_a"),
    ], "2026-Q2")
    write_snapshot_meta(tmp_db, "2026-Q2", 1, 0, ["src_a"], [])
    carried = carry_forward_records(tmp_db, "2026-Q3", {"src_a"})
    assert carried == []


def test_carry_forward_uses_only_latest_prior_snapshot(tmp_db):
    # src_b appears only in the older snapshot; it should NOT be resurrected
    # from an old snapshot once it has dropped out of the latest one.
    write_businesses(tmp_db, [
        _make_record(business_id="uuid-1", source_id="src_a"),
        _make_record(business_id="uuid-2", source_id="src_b"),
    ], "2026-Q1")
    write_snapshot_meta(tmp_db, "2026-Q1", 2, 0, ["src_a", "src_b"], [])
    write_businesses(tmp_db, [
        _make_record(business_id="uuid-1", source_id="src_a"),
    ], "2026-Q2")
    write_snapshot_meta(tmp_db, "2026-Q2", 1, 0, ["src_a"], [])

    carried = carry_forward_records(tmp_db, "2026-Q3", set())
    ids = {r["business_id"] for r in carried}
    assert ids == {"uuid-1"}   # only what was in the latest (2026-Q2) snapshot


# ── entity_resolver tests ────────────────────────────────────────────────────

from pipeline.entity_resolver import normalize_name, normalize_zip, resolve


def test_normalize_name_lowercases():
    assert normalize_name("ACME LLC") == "acme"


def test_normalize_name_strips_legal_suffixes():
    assert normalize_name("Smith Corp.") == "smith"
    assert normalize_name("Jones Inc") == "jones"
    assert normalize_name("Apex Enterprises") == "apex"


def test_normalize_name_strips_punctuation():
    assert normalize_name("A & B Services, LLC") == "a b"


def test_normalize_zip_pads_to_five():
    assert normalize_zip("1234") == "01234"


def test_normalize_zip_truncates_plus_four():
    assert normalize_zip("11201-1234") == "11201"


def _rec(name, zip_code, src_biz_id="", source_id="src_a"):
    return {
        "business_name": name,
        "address_zip": zip_code,
        "source_business_id": src_biz_id,
        "source_id": source_id,
    }


def test_resolve_assigns_new_uuid_when_no_match():
    records = [_rec("Brand New Biz", "10001")]
    registry = []
    review_log = []
    result, new_entries = resolve(records, registry, "2026-Q2", review_log)
    assert len(result) == 1
    assert len(result[0]["business_id"]) == 36  # UUID format
    assert len(new_entries) == 1


def test_resolve_matches_by_source_business_id():
    existing = [{"business_id": "existing-uuid", "canonical_name": "acme",
                 "canonical_zip": "10001", "source_id": "src_a",
                 "source_business_id": "ACC001", "first_seen": "2026-Q1",
                 "last_seen": "2026-Q1"}]
    records = [_rec("ACME LLC", "10001", src_biz_id="ACC001")]
    review_log = []
    result, new_entries = resolve(records, existing, "2026-Q2", review_log)
    assert result[0]["business_id"] == "existing-uuid"
    assert len(new_entries) == 0


def test_resolve_matches_by_name_and_zip():
    existing = [{"business_id": "existing-uuid", "canonical_name": "acme",
                 "canonical_zip": "10001", "source_id": "src_a",
                 "source_business_id": "", "first_seen": "2026-Q1",
                 "last_seen": "2026-Q1"}]
    records = [_rec("Acme LLC", "10001")]
    review_log = []
    result, new_entries = resolve(records, existing, "2026-Q2", review_log)
    assert result[0]["business_id"] == "existing-uuid"
    assert len(new_entries) == 0


def test_resolve_logs_uncertain_match():
    # "sunrise bakery" and "sunrise bakeries" normalize to themselves and score ~86.7%
    # — squarely in the 80-94% near-miss range that must be logged but not matched.
    existing = [{"business_id": "existing-uuid", "canonical_name": "sunrise bakery",
                 "canonical_zip": "10001", "source_id": "src_a",
                 "source_business_id": "", "first_seen": "2026-Q1",
                 "last_seen": "2026-Q1"}]
    records = [_rec("Sunrise Bakeries", "10001")]
    review_log = []
    result, new_entries = resolve(records, existing, "2026-Q2", review_log)
    assert len(review_log) == 1
    assert review_log[0]["candidate_id"] == "existing-uuid"
    assert result[0]["business_id"] != "existing-uuid"


def test_resolve_different_source_ids_dont_cross_match():
    existing = [{"business_id": "existing-uuid", "canonical_name": "acme",
                 "canonical_zip": "10001", "source_id": "src_a",
                 "source_business_id": "ACC001", "first_seen": "2026-Q1",
                 "last_seen": "2026-Q1"}]
    # Same source_business_id but different source — should not match
    records = [_rec("Acme LLC", "10001", src_biz_id="ACC001", source_id="src_b")]
    review_log = []
    result, new_entries = resolve(records, existing, "2026-Q2", review_log)
    assert result[0]["business_id"] != "existing-uuid"


def test_resolve_skips_records_with_missing_source_id():
    records = [{"business_name": "Some Biz", "address_zip": "10001",
                "source_business_id": "", "source_id": ""}]
    review_log = []
    import warnings
    with warnings.catch_warnings(record=True) as w:
        warnings.simplefilter("always")
        result, new_entries = resolve(records, [], "2026-Q2", review_log)
    assert len(result) == 0
    assert len(new_entries) == 0
    assert len(w) == 1


# ── geocoder tests ───────────────────────────────────────────────────────────

from unittest.mock import patch, MagicMock
from pipeline.geocoder import batch_geocode


def _make_census_response(rows: list[str]) -> MagicMock:
    """rows: list of CSV lines as the Census API would return."""
    mock = MagicMock()
    mock.raise_for_status.return_value = None
    mock.text = "\n".join(rows)
    return mock


def test_batch_geocode_returns_coords_for_matched_records():
    census_csv = [
        '"uuid-1","123 Main St, Brooklyn, NY, 11201","Match","Exact","123 Main St, Brooklyn, NY 11201","-73.944,40.678",1234567,L'
    ]
    with patch("requests.post", return_value=_make_census_response(census_csv)):
        result = batch_geocode([{
            "business_id": "uuid-1",
            "address_street": "123 Main St",
            "address_city": "Brooklyn",
            "address_state": "New York",
            "address_zip": "11201",
        }])
    assert "uuid-1" in result
    lat, lon = result["uuid-1"]
    assert abs(lat - 40.678) < 0.001
    assert abs(lon - (-73.944)) < 0.001


def test_batch_geocode_survives_network_failure():
    # A network drop (e.g. laptop sleeps) must NOT crash the pipeline and lose
    # the whole run — geocoding degrades to "no coords this run".
    import requests as _requests
    with patch("requests.post",
               side_effect=_requests.exceptions.ConnectionError("DNS fail")), \
         patch("time.sleep"):
        result = batch_geocode([{
            "business_id": "uuid-1", "address_street": "1 Main St",
            "address_city": "Hartford", "address_state": "CT",
            "address_zip": "06103",
        }])
    assert result == {}


def test_batch_geocode_skips_non_match():
    census_csv = [
        '"uuid-2","Bad Address, Nowhere, NY, 00000","No_Match","","","",,',
    ]
    with patch("requests.post", return_value=_make_census_response(census_csv)):
        result = batch_geocode([{
            "business_id": "uuid-2",
            "address_street": "Bad Address",
            "address_city": "Nowhere",
            "address_state": "NY",
            "address_zip": "00000",
        }])
    assert "uuid-2" not in result


def test_batch_geocode_returns_empty_dict_for_empty_input():
    with patch("requests.post") as mock_post:
        result = batch_geocode([])
    assert result == {}
    mock_post.assert_not_called()


def test_batch_geocode_filters_records_missing_coords():
    records = [
        {"business_id": "has-coords", "latitude": "40.68", "longitude": "-73.94",
         "address_street": "1 Main", "address_city": "Brooklyn",
         "address_state": "NY", "address_zip": "11201"},
        {"business_id": "no-coords", "latitude": "", "longitude": "",
         "address_street": "2 Main", "address_city": "Brooklyn",
         "address_state": "NY", "address_zip": "11201"},
    ]
    census_csv = [
        '"no-coords","2 Main, Brooklyn, NY, 11201","Match","Exact","2 Main, Brooklyn, NY 11201","-73.945,40.679",1234568,L'
    ]
    # Only no-coords should be submitted; has-coords should be skipped
    with patch("requests.post", return_value=_make_census_response(census_csv)) as mock_post:
        result = batch_geocode(records)
    call_args = mock_post.call_args
    submitted_csv = call_args.kwargs["files"]["addressFile"][1]
    assert "no-coords" in submitted_csv
    assert "has-coords" not in submitted_csv
    assert "no-coords" in result
    lat, lon = result["no-coords"]
    assert abs(lat - 40.679) < 0.001


def test_batch_geocode_handles_malformed_response_row():
    census_csv = [
        '"uuid-3","123 Main St, Brooklyn, NY, 11201","Match","Exact","123 Main St","-73.944,not-a-number",1234567,L',
        '"uuid-4","456 Oak Ave, Brooklyn, NY, 11201","Match","Exact","456 Oak Ave, Brooklyn","-74.001,40.700",1234568,L',
    ]
    with patch("requests.post", return_value=_make_census_response(census_csv)):
        result = batch_geocode([
            {"business_id": "uuid-3", "address_street": "123 Main St",
             "address_city": "Brooklyn", "address_state": "NY", "address_zip": "11201"},
            {"business_id": "uuid-4", "address_street": "456 Oak Ave",
             "address_city": "Brooklyn", "address_state": "NY", "address_zip": "11201"},
        ])
    assert "uuid-3" not in result  # malformed coords should be skipped
    assert "uuid-4" in result      # valid row should still be processed


# ── export tests ─────────────────────────────────────────────────────────────

import csv as csv_module
from pipeline.export import export_csv, write_summary
from pipeline.db import open_db, write_businesses, write_snapshot_meta


@pytest.fixture
def db_with_snapshot(tmp_path):
    db_path = tmp_path / "test.duckdb"
    con = open_db(db_path)
    records = [
        {**_make_record(business_id="biz-1", business_name="Alpha Biz",
                        confidence="confirmed_black", source_id="src_a")},
        {**_make_record(business_id="biz-2", business_name="Beta Biz",
                        confidence="mbe_unverified", source_id="src_b")},
        # biz-1 also appears in src_b as mbe_unverified — confirmed should win
        {**_make_record(business_id="biz-1", business_name="Alpha Biz",
                        confidence="mbe_unverified", source_id="src_b")},
    ]
    write_businesses(con, records, "2026-Q2")
    write_snapshot_meta(con, "2026-Q2", 2, 0, ["src_a", "src_b"], [])
    return con


def test_export_csv_creates_file(db_with_snapshot, tmp_path):
    out = tmp_path / "businesses.csv"
    export_csv(db_with_snapshot, out, "2026-Q2")
    assert out.exists()


def test_export_csv_has_header_row(db_with_snapshot, tmp_path):
    out = tmp_path / "businesses.csv"
    export_csv(db_with_snapshot, out, "2026-Q2")
    with open(out) as f:
        reader = csv_module.DictReader(f)
        from pipeline.export import EXPORT_COLUMNS as _EC
        assert list(reader.fieldnames) == _EC


def test_export_csv_deduplicates_confirmed_wins(db_with_snapshot, tmp_path):
    out = tmp_path / "businesses.csv"
    export_csv(db_with_snapshot, out, "2026-Q2")
    with open(out) as f:
        rows = list(csv_module.DictReader(f))
    # biz-1 appears in both sources — confirmed_black should win
    biz1_rows = [r for r in rows if r["business_id"] == "biz-1"]
    assert len(biz1_rows) == 1
    assert biz1_rows[0]["confidence"] == "confirmed_black"


def test_export_csv_row_count_equals_unique_businesses(db_with_snapshot, tmp_path):
    out = tmp_path / "businesses.csv"
    export_csv(db_with_snapshot, out, "2026-Q2")
    with open(out) as f:
        rows = list(csv_module.DictReader(f))
    assert len(rows) == 2  # biz-1 and biz-2, deduplicated


def test_export_csv_creates_parent_directories(db_with_snapshot, tmp_path):
    out = tmp_path / "nested" / "dir" / "businesses.csv"
    export_csv(db_with_snapshot, out, "2026-Q2")
    assert out.exists()


def test_write_summary_creates_file(tmp_path):
    path = tmp_path / "2026-Q2-summary.txt"
    write_summary(path, "2026-Q2", 100, 5, ["src_a"], ["src_b"])
    assert path.exists()
    content = path.read_text()
    assert "2026-Q2" in content
    assert "100" in content
    assert "src_b" in content
    assert "5" in content  # records_dropped


# ── orchestrator tests ───────────────────────────────────────────────────────

from pipeline.run import current_snapshot_id, discover_adapters
from datetime import date


def test_current_snapshot_id_q1():
    assert current_snapshot_id(date(2026, 1, 15)) == "2026-Q1"

def test_current_snapshot_id_q2():
    assert current_snapshot_id(date(2026, 5, 1)) == "2026-Q2"

def test_current_snapshot_id_q3():
    assert current_snapshot_id(date(2026, 8, 31)) == "2026-Q3"

def test_current_snapshot_id_q4():
    assert current_snapshot_id(date(2026, 11, 1)) == "2026-Q4"

def test_current_snapshot_id_q1_boundary():
    assert current_snapshot_id(date(2026, 3, 31)) == "2026-Q1"
    assert current_snapshot_id(date(2026, 4, 1)) == "2026-Q2"

def test_current_snapshot_id_year_boundary():
    assert current_snapshot_id(date(2026, 12, 31)) == "2026-Q4"
    assert current_snapshot_id(date(2027, 1, 1)) == "2027-Q1"


def test_discover_adapters_finds_concrete_classes(tmp_path, monkeypatch):
    # Write a minimal valid adapter to a temp adapters directory
    adapters_dir = tmp_path / "adapters"
    adapters_dir.mkdir()
    (adapters_dir / "__init__.py").write_text("")
    (adapters_dir / "test_adapter.py").write_text("""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))
from pipeline.adapter_base import AdapterBase
class TestAdapter(AdapterBase):
    SOURCE_ID = 'test'; SOURCE_NAME = 'Test'
    PROGRAM = 'MWBE'; GEOGRAPHY = 'X'; CONFIDENCE = 'confirmed_black'
    FIELD_MAP = {}
    def fetch(self): return []
    def parse(self, raw): return []
""")
    monkeypatch.syspath_prepend(str(tmp_path))
    adapters = discover_adapters(adapters_dir)
    assert len(adapters) == 1
    assert adapters[0].SOURCE_ID == "test"


# ── nyc_mwbe adapter tests ───────────────────────────────────────────────────

import openpyxl
from adapters.nyc_mwbe import NycMwbeAdapter


@pytest.fixture
def nyc_xlsx(tmp_path):
    """Minimal NYC MWBE xlsx — same format as the existing sample_xlsx fixture."""
    filepath = tmp_path / "nyc_mwbe.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Export Date:", "09/09/2025"])
    ws.append(["Matching Records:", 2])
    ws.append(["Search Parameters"])
    ws.append(["codecategory", "both"])
    ws.append([None])
    ws.append([
        "Account Number", "Vendor Formal Name", "Vendor DBA",
        "First Name", "Last Name", "Telephone", "Email",
        "Business Description", "Certification", "Certification Renewal Date",
        "Ethnicity", "Address Line 1", "Address Line 2", "City", "State", "Zip",
        "Mailing Address Line 1", "Mailing Address Line 2", "Mailing City",
        "Mailing State", "Mailing Zip", "Website", "Date of Establishment",
        "Aggregate Bonding Limit", "Signatory to Union Contract(s)",
        "6 digit NAICS code", "NAICS Sector", "NAICS Subsector", "NAICS Title",
        "Types of Construction Projects Performed", "NIGP codes",
        "Largest Value of Contract"
    ])
    ws.append([
        "ACC001", "Horizon Consulting LLC", "", "Jane", "Doe", "212-555-0001",
        "jane@horizon.com", "Management consulting services.", "MBE", "2026-01-01",
        "Black", "123 Main St", "", "Brooklyn", "New York", "11201",
        "", "", "", "", "", "https://horizon.com", "2015",
        "", "", "561110", "Services", "Administrative", "Management Consulting",
        "", "", ""
    ])
    ws.append([
        "ACC002", "BuildRight Inc", "", "Marcus", "Johnson", "718-555-0002",
        "", "General contractor.", "M/WBE", "2026-01-01",
        "Black", "456 Atlantic Ave", "", "Bronx", "New York", "10451",
        "", "", "", "", "", "", None,
        "", "", "236220", "Construction", "Building Construction", "Commercial",
        "", "", ""
    ])
    ws.append([
        "ACC003", "Other Corp", "", "Ana", "Lopez", "", "",
        "Non-black business.", "MBE", "2026-01-01",
        "Hispanic", "789 Broadway", "", "Manhattan", "New York", "10013",
        "", "", "", "", "", "", "2020",
        "", "", "541511", "Technology", "Software", "Custom Software",
        "", "", ""
    ])
    wb.save(filepath)
    return filepath


def test_nyc_adapter_metadata():
    adapter = NycMwbeAdapter()
    assert adapter.SOURCE_ID == "nyc_mwbe"
    assert adapter.CONFIDENCE == "confirmed_black"
    assert adapter.PROGRAM == "MWBE"
    assert adapter.GEOGRAPHY == "NYC"


def test_nyc_adapter_filters_to_black_only(nyc_xlsx):
    adapter = NycMwbeAdapter(source_file=nyc_xlsx)
    records = adapter.run()
    assert len(records) == 2
    names = [r["business_name"] for r in records]
    assert "Other Corp" not in names


def test_nyc_adapter_maps_standard_fields(nyc_xlsx):
    adapter = NycMwbeAdapter(source_file=nyc_xlsx)
    records = adapter.run()
    rec = next(r for r in records if r["business_name"] == "Horizon Consulting LLC")
    assert rec["owner_name"] == "Jane Doe"
    assert rec["address_street"] == "123 Main St"
    assert rec["address_city"] == "Brooklyn"
    assert rec["address_state"] == "New York"
    assert rec["address_zip"] == "11201"
    assert rec["industry"] == "Services"
    assert rec["naics_code"] == "561110"
    assert rec["certification"] == "MBE"
    assert rec["website"] == "https://horizon.com"
    assert rec["year_founded"] == "2015"
    assert rec["phone"] == "212-555-0001"
    assert rec["email"] == "jane@horizon.com"
    assert rec["source_business_id"] == "ACC001"


def test_nyc_adapter_puts_extra_columns_in_source_fields(nyc_xlsx):
    adapter = NycMwbeAdapter(source_file=nyc_xlsx)
    records = adapter.run()
    rec = records[0]
    sf = rec["source_fields"]
    # Columns not in FIELD_MAP should land in source_fields
    assert "Vendor DBA" in sf


def test_nyc_adapter_handles_missing_year(nyc_xlsx):
    adapter = NycMwbeAdapter(source_file=nyc_xlsx)
    records = adapter.run()
    rec = next(r for r in records if r["business_name"] == "BuildRight Inc")
    assert rec["year_founded"] == ""


def test_nyc_adapter_sets_last_verified(nyc_xlsx):
    adapter = NycMwbeAdapter(source_file=nyc_xlsx)
    records = adapter.run()
    assert records[0]["last_verified"] == "2025-09-09"


def test_nyc_adapter_skips_none_ethnicity_rows(nyc_xlsx, tmp_path):
    # Add a row with None Ethnicity to the existing fixture
    import openpyxl
    wb = openpyxl.load_workbook(nyc_xlsx)
    ws = wb.active
    # Append a row where Ethnicity (col index 10, 0-based) is None
    row_data = ["ACC004", "Mystery Corp", "", "Unknown", "Person", "", "",
                "Unknown business.", "MBE", "2026-01-01",
                None,  # Ethnicity is None
                "100 Unknown St", "", "Queens", "New York", "11415",
                "", "", "", "", "", "", "2010", "", "", "541511",
                "Technology", "Software", "Custom", "", "", ""]
    ws.append(row_data)
    modified_path = tmp_path / "modified.xlsx"
    wb.save(modified_path)

    adapter = NycMwbeAdapter(source_file=modified_path)
    records = adapter.run()
    names = [r["business_name"] for r in records]
    assert "Mystery Corp" not in names


# ── sam_8a adapter tests ─────────────────────────────────────────────────────

import os
from unittest.mock import patch, MagicMock
from adapters.sam_8a import SamEightAAdapter


def _make_sam_response(entities: list[dict], total: int) -> MagicMock:
    mock = MagicMock()
    mock.raise_for_status.return_value = None
    mock.json.return_value = {"totalRecords": total, "entityData": entities}
    return mock


def _make_entity(uei="UEI001", name="Acme LLC", street="123 Main St",
                 city="Atlanta", state="GA", zipcode="30301",
                 url="https://acme.com", naics="541611"):
    return {
        "entityRegistration": {"ueiSAM": uei, "legalBusinessName": name},
        "coreData": {
            "physicalAddress": {
                "addressLine1": street,
                "city": city,
                "stateOrProvinceCode": state,
                "zipCode": zipcode,
            },
            "entityInformation": {"entityURL": url},
        },
        "assertions": {"goodsAndServices": {"primaryNaics": naics}},
    }


def test_sam_adapter_metadata():
    adapter = SamEightAAdapter(api_key="test-key")
    assert adapter.SOURCE_ID == "sam_8a"
    assert adapter.CONFIDENCE == "mbe_unverified"
    assert adapter.PROGRAM == "8(a)"
    assert adapter.GEOGRAPHY == "National"


def test_sam_adapter_raises_without_api_key(monkeypatch):
    monkeypatch.delenv("SAM_GOV_API_KEY", raising=False)
    with pytest.raises(ValueError, match="SAM_GOV_API_KEY"):
        SamEightAAdapter()


def test_sam_adapter_uses_env_api_key(monkeypatch):
    monkeypatch.setenv("SAM_GOV_API_KEY", "env-key-123")
    adapter = SamEightAAdapter()
    assert adapter._api_key == "env-key-123"


def test_sam_adapter_paginates_all_pages():
    entity = _make_entity()
    # totalRecords=15 → 2 pages (10 + 5)
    responses = [
        _make_sam_response([entity] * 10, total=15),
        _make_sam_response([entity] * 5, total=15),
    ]
    with patch("requests.get", side_effect=responses):
        raw = SamEightAAdapter(api_key="k").fetch()
    assert len(raw) == 15


def test_sam_adapter_maps_standard_fields():
    entity = _make_entity(
        uei="UEI999", name="Horizon Consulting LLC",
        street="123 Main St", city="Atlanta", state="GA",
        zipcode="30301", url="https://horizon.com", naics="541611",
    )
    with patch("requests.get", return_value=_make_sam_response([entity], 1)):
        records = SamEightAAdapter(api_key="k").run()
    rec = records[0]
    assert rec["business_name"] == "Horizon Consulting LLC"
    assert rec["address_street"] == "123 Main St"
    assert rec["address_city"] == "Atlanta"
    assert rec["address_state"] == "GA"
    assert rec["address_zip"] == "30301"
    assert rec["website"] == "https://horizon.com"
    assert rec["naics_code"] == "541611"
    assert rec["certification"] == "8(a)"


def test_sam_adapter_sets_uei_as_source_business_id():
    entity = _make_entity(uei="MYUEI123")
    with patch("requests.get", return_value=_make_sam_response([entity], 1)):
        records = SamEightAAdapter(api_key="k").run()
    assert records[0]["source_business_id"] == "MYUEI123"


def test_sam_adapter_puts_extra_fields_in_source_fields():
    entity = _make_entity()
    # The adapter flattens nested SAM data; anything beyond FIELD_MAP lands in source_fields
    with patch("requests.get", return_value=_make_sam_response([entity], 1)):
        records = SamEightAAdapter(api_key="k").run()
    assert "source_fields" in records[0]
    assert isinstance(records[0]["source_fields"], dict)


def test_sam_adapter_handles_missing_optional_field():
    # entityURL missing → website should be ""
    entity = _make_entity()
    del entity["coreData"]["entityInformation"]
    with patch("requests.get", return_value=_make_sam_response([entity], 1)):
        records = SamEightAAdapter(api_key="k").run()
    assert records[0]["website"] == ""


def test_sam_adapter_returns_empty_on_zero_results():
    with patch("requests.get", return_value=_make_sam_response([], 0)):
        records = SamEightAAdapter(api_key="k").run()
    assert records == []


# ── tx_hub adapter tests ──────────────────────────────────────────────────────

import csv
import io
from adapters.tx_hub import TxHubAdapter


def _make_hub_csv(rows: list[dict]) -> str:
    """Build a CSV string in Texas HUB format from a list of dicts."""
    fieldnames = [
        "VENDOR ID NUMBER", " VENDOR NAME", " VENDOR ADDRESS LINE 1", "VENDOR ADDRESS LINE 2",
        "CITY", "STATE", "ZIP CODE", " FOREIGN ADDRESS", "PHONE NUMBER", " FAX NUMBER",
        "GENDER", "ELIGIBILITY CODE", " STATUS CODE", "COUNTY", "BUSINESS DESCRIPTION",
        " VENDOR NUMBER", "EXPIRATION DATE", " CONTACT NAME", "TEXAS OFFICE FLAG",
        "INTERNET ADDRESS", " QISV FLAG", "SDV FLAG", " SMALL BUSINESS FLAG",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def _make_hub_row(vendor_id="1000000000001", name="Acme Black LLC",
                  street="100 Main St", city="Houston", state="TX",
                  zipcode="77001", phone="713-555-0100", website="https://acme.com",
                  eligibility="BL", status="D", description="Consulting services"):
    return {
        "VENDOR ID NUMBER": vendor_id,
        " VENDOR NAME": name,
        " VENDOR ADDRESS LINE 1": street,
        "VENDOR ADDRESS LINE 2": "",
        "CITY": city,
        "STATE": state,
        "ZIP CODE": zipcode,
        " FOREIGN ADDRESS": "USA",
        "PHONE NUMBER": phone,
        " FAX NUMBER": "",
        "GENDER": "M",
        "ELIGIBILITY CODE": eligibility,
        " STATUS CODE": status,
        "COUNTY": "HARRIS",
        "BUSINESS DESCRIPTION": description,
        " VENDOR NUMBER": "123456",
        "EXPIRATION DATE": "05-JAN-2026",
        " CONTACT NAME": "Jane Smith",
        "TEXAS OFFICE FLAG": "Y",
        "INTERNET ADDRESS": website,
        " QISV FLAG": "",
        "SDV FLAG": "",
        " SMALL BUSINESS FLAG": "Y",
    }


def _mock_hub_response(rows: list[dict]) -> MagicMock:
    mock = MagicMock()
    mock.raise_for_status.return_value = None
    mock.text = _make_hub_csv(rows)
    return mock


def test_tx_hub_adapter_metadata():
    adapter = TxHubAdapter()
    assert adapter.SOURCE_ID == "tx_hub"
    assert adapter.CONFIDENCE == "confirmed_black"
    assert adapter.PROGRAM == "HUB"
    assert adapter.GEOGRAPHY == "Texas"


def test_tx_hub_filters_bl_only():
    rows = [
        _make_hub_row(eligibility="BL", name="Black Firm"),
        _make_hub_row(eligibility="HI", name="Hispanic Firm"),
        _make_hub_row(eligibility="AS", name="Asian Firm"),
    ]
    with patch("requests.get", return_value=_mock_hub_response(rows)):
        raw = TxHubAdapter().fetch()
    assert len(raw) == 1
    assert raw[0]["VENDOR NAME"] == "Black Firm"


def test_tx_hub_maps_standard_fields():
    row = _make_hub_row(
        name="Houston Consulting LLC", street="100 Main St", city="Houston",
        state="TX", zipcode="77001", phone="713-555-0100", website="https://hc.com",
        description="Management consulting",
    )
    with patch("requests.get", return_value=_mock_hub_response([row])):
        records = TxHubAdapter().run()
    rec = records[0]
    assert rec["business_name"] == "Houston Consulting LLC"
    assert rec["address_street"] == "100 Main St"
    assert rec["address_city"] == "Houston"
    assert rec["address_state"] == "TX"
    assert rec["address_zip"] == "77001"
    assert rec["phone"] == "713-555-0100"
    assert rec["website"] == "https://hc.com"
    assert rec["description"] == "Management consulting"
    assert rec["certification"] == "HUB"


def test_tx_hub_sets_vendor_id_as_source_business_id():
    row = _make_hub_row(vendor_id="9876543210001")
    with patch("requests.get", return_value=_mock_hub_response([row])):
        records = TxHubAdapter().run()
    assert records[0]["source_business_id"] == "9876543210001"


def test_tx_hub_preserves_status_in_source_fields():
    row = _make_hub_row(status="D")
    with patch("requests.get", return_value=_mock_hub_response([row])):
        records = TxHubAdapter().run()
    assert "source_fields" in records[0]
    assert records[0]["source_fields"].get("STATUS CODE") == "D"


def test_tx_hub_returns_empty_on_no_bl_records():
    rows = [_make_hub_row(eligibility="HI"), _make_hub_row(eligibility="WO")]
    with patch("requests.get", return_value=_mock_hub_response(rows)):
        records = TxHubAdapter().run()
    assert records == []


def test_tx_hub_strips_column_whitespace():
    # Columns like ' VENDOR NAME' and ' STATUS CODE' have leading spaces in the raw CSV
    row = _make_hub_row(name="  Spaced Name  ")
    with patch("requests.get", return_value=_mock_hub_response([row])):
        raw = TxHubAdapter().fetch()
    # After stripping, the key should be clean and value accessible
    assert "VENDOR NAME" in raw[0]


# ── md_mbe adapter tests ──────────────────────────────────────────────────────

import tempfile
from adapters.md_mbe import MdMbeAdapter


def _make_md_csv(rows: list[dict]) -> str:
    """Build a CSV string in the real MDOT/gob2g export format:
    5 metadata preamble rows, then the header row, then data rows.
    (The adapter skips METADATA_ROWS=5 lines before the header.)"""
    fieldnames = [
        "Company Name", "Physical Address", "City", "State", "Zip", "Phone",
        "Email", "Website", "Minority Status", "Certification Number",
        "Owner First", "Owner Last",
    ]
    buf = io.StringIO()
    for i in range(5):
        buf.write(f"MDOT MBE Directory export metadata row {i + 1}\n")
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def _make_md_row(cert_number="MD001", name="Baltimore Tech LLC",
                 street="200 Light St", city="Baltimore", state="MD",
                 zipcode="21202", phone="410-555-0200", email="info@baltech.com",
                 website="https://baltech.com", minority_status="African American",
                 owner_first="John", owner_last="Doe"):
    return {
        "Company Name": name,
        "Physical Address": street,
        "City": city,
        "State": state,
        "Zip": zipcode,
        "Phone": phone,
        "Email": email,
        "Website": website,
        "Minority Status": minority_status,
        "Certification Number": cert_number,
        "Owner First": owner_first,
        "Owner Last": owner_last,
    }


def test_md_mbe_adapter_metadata(tmp_path):
    csv_file = tmp_path / "md_mbe.csv"
    csv_file.write_text(_make_md_csv([_make_md_row()]))
    adapter = MdMbeAdapter(file_path=csv_file)
    assert adapter.SOURCE_ID == "md_mbe"
    assert adapter.CONFIDENCE == "confirmed_black"
    assert adapter.PROGRAM == "MBE"
    assert adapter.GEOGRAPHY == "Maryland"


def test_md_mbe_raises_without_file(monkeypatch):
    monkeypatch.delenv("MD_MBE_FILE", raising=False)
    with pytest.raises(ValueError, match="MD_MBE_FILE"):
        MdMbeAdapter()


def test_md_mbe_uses_env_file_path(monkeypatch, tmp_path):
    csv_file = tmp_path / "md_mbe.csv"
    csv_file.write_text(_make_md_csv([_make_md_row()]))
    monkeypatch.setenv("MD_MBE_FILE", str(csv_file))
    adapter = MdMbeAdapter()
    assert adapter._file_path == csv_file


def test_md_mbe_maps_standard_fields(tmp_path):
    row = _make_md_row(
        name="Baltimore Tech LLC", street="200 Light St", city="Baltimore",
        state="MD", zipcode="21202", phone="410-555-0200",
        email="info@baltech.com", website="https://baltech.com",
    )
    csv_file = tmp_path / "md_mbe.csv"
    csv_file.write_text(_make_md_csv([row]))
    records = MdMbeAdapter(file_path=csv_file).run()
    rec = records[0]
    assert rec["business_name"] == "Baltimore Tech LLC"
    assert rec["address_street"] == "200 Light St"
    assert rec["address_city"] == "Baltimore"
    assert rec["address_state"] == "MD"
    assert rec["address_zip"] == "21202"
    assert rec["phone"] == "410-555-0200"
    assert rec["email"] == "info@baltech.com"
    assert rec["website"] == "https://baltech.com"
    assert rec["certification"] == "MBE"


def test_md_mbe_sets_cert_number_as_source_business_id(tmp_path):
    row = _make_md_row(cert_number="MD99999")
    csv_file = tmp_path / "md_mbe.csv"
    csv_file.write_text(_make_md_csv([row]))
    records = MdMbeAdapter(file_path=csv_file).run()
    assert records[0]["source_business_id"] == "MD99999"


def test_md_mbe_puts_extra_fields_in_source_fields(tmp_path):
    row = _make_md_row()
    csv_file = tmp_path / "md_mbe.csv"
    csv_file.write_text(_make_md_csv([row]))
    records = MdMbeAdapter(file_path=csv_file).run()
    assert "source_fields" in records[0]
    assert isinstance(records[0]["source_fields"], dict)
    # Minority Status is not in FIELD_MAP → should be in source_fields
    assert "Minority Status" in records[0]["source_fields"]


def test_md_mbe_handles_empty_file(tmp_path):
    # Valid file (5 metadata rows + header) but zero data rows → no records.
    csv_file = tmp_path / "md_mbe.csv"
    csv_file.write_text(_make_md_csv([]))
    records = MdMbeAdapter(file_path=csv_file).run()
    assert records == []


def test_md_mbe_raises_file_not_found():
    with pytest.raises(FileNotFoundError):
        MdMbeAdapter(file_path=Path("/nonexistent/md_mbe.csv"))


# ── in_idoa adapter tests ─────────────────────────────────────────────────────

import openpyxl as _openpyxl
from adapters.in_idoa import InIdoaAdapter

_IN_HEADERS = [
    "Company Name", "DBA", "UNSPSC", "UNSPSC Description",
    "First Name", "LastName", "Mailing Address 1", "Mailing Address 2",
    "City", "State", "Zip Code", "County", "Application Type",
    "Ethnic Group", "Certification Date", "Expiration Dte",
    "Bidder ID", "Email ID", "Phone", "Application Status", "Company Name Upper",
]


def _make_in_xlsx(tmp_path, rows: list[list]) -> Path:
    """
    Build an Indiana IDOA-format xlsx fixture.
    Row 1: title row. Row 2: column headers. Row 3+: data.
    """
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.append(["Diversity Certified Businesses", len(rows)] + [None] * 19)
    ws.append(_IN_HEADERS)
    for row in rows:
        ws.append(row)
    path = tmp_path / "in_idoa.xlsx"
    wb.save(path)
    return path


def _in_row(company="Acme Black LLC", dba="Acme Black LLC",
            unspsc="54111500", unspsc_desc="Office supplies",
            first="Jordan", last="Smith",
            street="100 N Meridian St", street2=None,
            city="Indianapolis", state="IN", zipcode="46204",
            county="Marion", app_type="MBE", ethnic="AFA",
            cert_date=None, exp_date=None,
            bidder_id="0000001234", email="jordan@acme.com",
            phone="317/555-0100", status="CERT", name_upper="ACME BLACK LLC"):
    return [
        company, dba, unspsc, unspsc_desc, first, last,
        street, street2, city, state, zipcode, county,
        app_type, ethnic, cert_date, exp_date,
        bidder_id, email, phone, status, name_upper,
    ]


@pytest.fixture()
def in_xlsx(tmp_path):
    rows = [
        _in_row(company="Acme Black LLC", bidder_id="0000001234",
                first="Jordan", last="Smith", ethnic="AFA"),
        # Same firm, second UNSPSC code — should be deduplicated away
        _in_row(company="Acme Black LLC", bidder_id="0000001234",
                unspsc="54111600", unspsc_desc="Paper products",
                first="Jordan", last="Smith", ethnic="AFA"),
        # Different AFA firm
        _in_row(company="BuildRight Inc", bidder_id="0000005678",
                first="Alex", last="Johnson",
                street="200 S Capitol Ave", city="Indianapolis",
                zipcode="46225", app_type="WBE", ethnic="AFA"),
        # Non-AFA firm — should be excluded
        _in_row(company="Other Corp", bidder_id="0000009999",
                ethnic="CAU"),
    ]
    return _make_in_xlsx(tmp_path, rows)


def test_in_idoa_metadata():
    adapter = InIdoaAdapter(file_path=SOURCE_FILE) if False else None
    # Instantiate without a real file for metadata checks
    import unittest.mock as _mock
    with _mock.patch.object(Path, "exists", return_value=True):
        adapter = InIdoaAdapter(file_path=Path("/fake/path.xlsx"))
    assert adapter.SOURCE_ID == "in_idoa"
    assert adapter.CONFIDENCE == "confirmed_black"
    assert adapter.PROGRAM == "MBE"
    assert adapter.GEOGRAPHY == "Indiana"


def test_in_idoa_filters_to_afa_only(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    names = [r["business_name"] for r in records]
    assert "Other Corp" not in names
    assert "Acme Black LLC" in names
    assert "BuildRight Inc" in names


def test_in_idoa_deduplicates_by_bidder_id(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    # Acme Black LLC appears twice in the file (two UNSPSC codes) — only one record
    acme_records = [r for r in records if r["business_name"] == "Acme Black LLC"]
    assert len(acme_records) == 1


def test_in_idoa_total_count(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    assert len(records) == 2


def test_in_idoa_maps_standard_fields(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    rec = next(r for r in records if r["business_name"] == "Acme Black LLC")
    assert rec["address_street"] == "100 N Meridian St"
    assert rec["address_city"] == "Indianapolis"
    assert rec["address_state"] == "IN"
    assert rec["address_zip"] == "46204"
    assert rec["email"] == "jordan@acme.com"
    assert rec["phone"] == "317/555-0100"


def test_in_idoa_owner_name(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    rec = next(r for r in records if r["business_name"] == "Acme Black LLC")
    assert rec["owner_name"] == "Jordan Smith"


def test_in_idoa_source_business_id(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    rec = next(r for r in records if r["business_name"] == "Acme Black LLC")
    assert rec["source_business_id"] == "0000001234"


def test_in_idoa_certification_reflects_app_type(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    mbe_rec = next(r for r in records if r["business_name"] == "Acme Black LLC")
    wbe_rec = next(r for r in records if r["business_name"] == "BuildRight Inc")
    assert mbe_rec["certification"] == "MBE"
    assert wbe_rec["certification"] == "WBE"


def test_in_idoa_sets_last_verified(in_xlsx):
    from datetime import date
    records = InIdoaAdapter(file_path=in_xlsx).run()
    assert records[0]["last_verified"] == str(date.today())


def test_in_idoa_extra_columns_in_source_fields(in_xlsx):
    records = InIdoaAdapter(file_path=in_xlsx).run()
    sf = records[0]["source_fields"]
    assert "Ethnic Group" in sf
    assert "County" in sf


def test_in_idoa_raises_file_not_found():
    with pytest.raises(FileNotFoundError):
        InIdoaAdapter(file_path=Path("/nonexistent/in_idoa.xlsx"))


# ── ct_das_smbe adapter tests ─────────────────────────────────────────────────

from adapters.ct_das_smbe import CtDasSmbeAdapter

_CT_HEADERS = [
    "vendorname", "business_address1", "townnamecrosswalk_standardized_town",
    "zip", "county", "business_state", "certification_type",
    "class_description_detailed", "active_date", "expiration_date", "status",
    "product", "gs_code", "goods_and_services", "location",
]


def _make_ct_csv(rows: list[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_CT_HEADERS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def _ct_row(vendorname="Down To Earth Consulting LLC",
            street="27 Siemon Company Drive", city="Watertown",
            state="CT", zipcode="06795", cert_type="MBE",
            ethnicity="Black American", gs_code="541330",
            goods="Engineering Services", product="Geotechnical Services",
            location="POINT (-73.11257 41.60364)"):
    return {
        "vendorname": vendorname, "business_address1": street,
        "townnamecrosswalk_standardized_town": city, "zip": zipcode,
        "county": "NA", "business_state": state,
        "certification_type": cert_type,
        "class_description_detailed": ethnicity,
        "active_date": "2028-05-22T00:00:00.000",
        "expiration_date": "2030-05-21T00:00:00.000", "status": "Certified",
        "product": product, "gs_code": gs_code, "goods_and_services": goods,
        "location": location,
    }


def _ct_response(csv_text: str):
    resp = MagicMock()
    resp.text = csv_text
    resp.raise_for_status = MagicMock()
    return resp


def test_ct_filters_to_black_american():
    csv_text = _make_ct_csv([
        _ct_row(vendorname="Black Co", ethnicity="Black American"),
        _ct_row(vendorname="Hisp Co", ethnicity="Hispanic American"),
        _ct_row(vendorname="Iberian Co", ethnicity="Iberian Peninsula"),
    ])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        records = CtDasSmbeAdapter().run()
    assert [r["business_name"] for r in records] == ["Black Co"]


def test_ct_maps_standard_fields():
    csv_text = _make_ct_csv([_ct_row(
        vendorname="Black Co", street="27 Siemon Company Drive",
        city="Watertown", state="CT", zipcode="06795", gs_code="541330",
    )])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        rec = CtDasSmbeAdapter().run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_street"] == "27 Siemon Company Drive"
    assert rec["address_city"] == "Watertown"
    assert rec["address_state"] == "CT"
    assert rec["address_zip"] == "06795"
    assert rec["naics_code"] == "541330"


def test_ct_certification_from_type():
    csv_text = _make_ct_csv([_ct_row(cert_type="SBE")])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        rec = CtDasSmbeAdapter().run()[0]
    assert rec["certification"] == "SBE"


def test_ct_extracts_coords_from_point():
    csv_text = _make_ct_csv([_ct_row(location="POINT (-73.11257 41.60364)")])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        rec = CtDasSmbeAdapter().run()[0]
    assert rec["latitude"] == "41.60364"
    assert rec["longitude"] == "-73.11257"


def test_ct_handles_missing_point():
    csv_text = _make_ct_csv([_ct_row(location="")])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        rec = CtDasSmbeAdapter().run()[0]
    assert rec["latitude"] == ""
    assert rec["longitude"] == ""


def test_ct_confidence_is_confirmed_black():
    assert CtDasSmbeAdapter.CONFIDENCE == "confirmed_black"


def test_ct_empty_when_no_black_american():
    csv_text = _make_ct_csv([_ct_row(ethnicity="Hispanic American")])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        records = CtDasSmbeAdapter().run()
    assert records == []


# ── de_osd adapter tests ──────────────────────────────────────────────────────

from adapters.de_osd import DeOsdAdapter

_DE_HEADERS = [
    "name", "certificatenumber", "primarycontactname", "address", "city",
    "state", "zipcode", "phonenumber", "email", "description",
    "ddd_baa", "ddd_ha", "ddd_f", "ct_mbe", "ct_wbe",
]


def _make_de_csv(rows: list[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_DE_HEADERS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def _de_row(name="1st State Aerials LLC", cert="DE24074827",
            contact="Gregory Morris", address="409 W. 19th Street",
            city="Wilmington", state="Delaware", zipcode="19802",
            phone="3024943627", email="x@y.com", description="Aerial services",
            baa="YES", ct_mbe="YES"):
    return {
        "name": name, "certificatenumber": cert, "primarycontactname": contact,
        "address": address, "city": city, "state": state, "zipcode": zipcode,
        "phonenumber": phone, "email": email, "description": description,
        "ddd_baa": baa, "ddd_ha": "", "ddd_f": "", "ct_mbe": ct_mbe, "ct_wbe": "",
    }


def test_de_filters_to_baa_yes():
    csv_text = _make_de_csv([
        _de_row(name="Black Co", baa="YES"),
        _de_row(name="NonBlack Co", baa=""),
        _de_row(name="Hisp Co", baa=" "),
    ])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        records = DeOsdAdapter().run()
    assert [r["business_name"] for r in records] == ["Black Co"]


def test_de_maps_standard_fields():
    csv_text = _make_de_csv([_de_row(
        name="Black Co", contact="Gregory Morris", address="409 W. 19th Street",
        city="Wilmington", state="Delaware", zipcode="19802",
        phone="3024943627", email="g@example.com",
    )])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        rec = DeOsdAdapter().run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["owner_name"] == "Gregory Morris"
    assert rec["address_street"] == "409 W. 19th Street"
    assert rec["address_city"] == "Wilmington"
    assert rec["address_state"] == "Delaware"
    assert rec["address_zip"] == "19802"
    assert rec["phone"] == "3024943627"
    assert rec["email"] == "g@example.com"


def test_de_source_business_id_is_cert_number():
    csv_text = _make_de_csv([_de_row(cert="DE99999")])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        rec = DeOsdAdapter().run()[0]
    assert rec["source_business_id"] == "DE99999"


def test_de_confidence_is_confirmed_black():
    assert DeOsdAdapter.CONFIDENCE == "confirmed_black"


def test_de_empty_when_no_baa():
    csv_text = _make_de_csv([_de_row(baa="")])
    with patch("requests.get", return_value=_ct_response(csv_text)):
        records = DeOsdAdapter().run()
    assert records == []


# ── sc_smbcc adapter tests ────────────────────────────────────────────────────

from adapters.sc_smbcc import ScSmbccAdapter

_SC_HEADERS = [
    None, "Organization Lookup", None, "DBA", "Business Address", "Business City",
    "Business State", "Business Zip", "Year Established", "Business Phone",
    "Vendor Registration Number", "Services", "Service Area", "Business Email",
    "Class", "Certification ID", "Date Certified", "Expiration Date",
]


def _make_sc_xlsx_bytes(data_rows: list[list]) -> bytes:
    """Build an SMBCC-format workbook: preamble rows, a header row, then data."""
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.append([None])
    ws.append([None, "SMBCC Excel Report for Website"])
    ws.append([None, "As of 2026-06-02"])
    for _ in range(6):
        ws.append([None])              # variable preamble (header is NOT at a fixed row)
    ws.append(_SC_HEADERS)
    for row in data_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sc_data_row(name="Palmetto Builders LLC", dba="", address="100 Main St",
                 city="Columbia", state="SC", zipcode="29201", year="2015",
                 phone="803-555-0100", vendor_id="V12345", services="Construction",
                 area="Statewide", email="info@palmetto.com",
                 class_val="01 - African American Male Owners",
                 cert_id="202581", date_cert="2024-01-01", exp="2026-01-01"):
    # Positions match _SC_HEADERS (index 0 and 2 are unnamed/blank columns).
    return [None, name, None, dba, address, city, state, zipcode, year, phone,
            vendor_id, services, area, email, class_val, cert_id, date_cert, exp]


def _sc_fetch_mock(xlsx_bytes: bytes):
    """side_effect for requests.get: 1st call = landing HTML, 2nd = xlsx bytes."""
    landing = MagicMock()
    landing.text = '<a href="/files/SC%20Cert%20List%2006.02.26.xlsx">download</a>'
    landing.raise_for_status = MagicMock()
    xlsx = MagicMock()
    xlsx.content = xlsx_bytes
    xlsx.raise_for_status = MagicMock()
    return [landing, xlsx]


def test_sc_filters_to_black_class_codes():
    xlsx = _make_sc_xlsx_bytes([
        _sc_data_row(name="AA Male Co", class_val="01 - African American Male Owners"),
        _sc_data_row(name="AA Female Co", class_val="02 - African American Female Owners"),
        _sc_data_row(name="DLT AA Co", class_val="05 - DLT Certified AA Male/Female"),
        _sc_data_row(name="Cauc Female Co", class_val="03- Caucasian Female Owners"),
        _sc_data_row(name="Hisp Co", class_val="04 - Hispanic Male/Female Owners"),
        _sc_data_row(name="Asian Co", class_val="09 - Asian Pacific or Other"),
    ])
    with patch("requests.get", side_effect=_sc_fetch_mock(xlsx)):
        records = ScSmbccAdapter().run()
    names = sorted(r["business_name"] for r in records)
    assert names == ["AA Female Co", "AA Male Co", "DLT AA Co"]


def test_sc_maps_standard_fields():
    xlsx = _make_sc_xlsx_bytes([_sc_data_row(
        name="Palmetto Builders LLC", address="100 Main St", city="Columbia",
        state="SC", zipcode="29201", phone="803-555-0100", email="info@palmetto.com",
        vendor_id="V12345",
    )])
    with patch("requests.get", side_effect=_sc_fetch_mock(xlsx)):
        rec = ScSmbccAdapter().run()[0]
    assert rec["business_name"] == "Palmetto Builders LLC"
    assert rec["address_street"] == "100 Main St"
    assert rec["address_city"] == "Columbia"
    assert rec["address_state"] == "SC"
    assert rec["address_zip"] == "29201"
    assert rec["phone"] == "803-555-0100"
    assert rec["email"] == "info@palmetto.com"
    assert rec["source_business_id"] == "V12345"


def test_sc_confidence_is_confirmed_black():
    assert ScSmbccAdapter.CONFIDENCE == "confirmed_black"


def test_sc_empty_when_no_black():
    xlsx = _make_sc_xlsx_bytes([
        _sc_data_row(class_val="03- Caucasian Female Owners"),
    ])
    with patch("requests.get", side_effect=_sc_fetch_mock(xlsx)):
        records = ScSmbccAdapter().run()
    assert records == []


# ── or_cobid adapter tests ────────────────────────────────────────────────────

from adapters.or_cobid import OrCobidAdapter

_OR_HEADERS = ["Company Name", "DBA Name", "Owner First", "Owner Last", "Location",
               "Phone", "Email", "Website", "Agency", "Certification Type",
               "Ethnicity", "Gender", "Capability", "County"]


def _make_dir_csv(title, headers, rows):
    """Build a gob2g/dbesystem-style export: title preamble, header row, data."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([title]); w.writerow(["As of 6/11/2026"]); w.writerow(["Results filtered"])
    w.writerow([]); w.writerow(["The information provided..."]); w.writerow([])
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    return buf.getvalue()


def _or_row(name="Acme Black LLC", first="Jordan", last="Smith",
            location="Portland, OR", phone="503-555-0100", email="a@b.com",
            website="https://acme.com", cert="MBE",
            ethnicity="African American (Black)", capability="Consulting"):
    return [name, "", first, last, location, phone, email, website, "Oregon",
            cert, ethnicity, "Male", capability, "Multnomah"]


def _write_latin1(tmp_path, name, text):
    p = tmp_path / name
    p.write_text(text, encoding="latin-1")
    return p


def test_or_filters_to_african_american(tmp_path):
    csv_text = _make_dir_csv("COBID Certified Firms Directory", _OR_HEADERS, [
        _or_row(name="Black Co", ethnicity="African American (Black)"),
        _or_row(name="White Co", ethnicity="Caucasian (White)"),
        _or_row(name="Hisp Co", ethnicity="Hispanic"),
    ])
    p = _write_latin1(tmp_path, "Oregon Directory test.csv", csv_text)
    records = OrCobidAdapter(file_path=p).run()
    assert [r["business_name"] for r in records] == ["Black Co"]


def test_or_maps_fields_and_parses_location(tmp_path):
    csv_text = _make_dir_csv("COBID Certified Firms Directory", _OR_HEADERS, [
        _or_row(name="Black Co", first="Jordan", last="Smith",
                location="Portland, OR", phone="503-555-0100",
                email="j@black.co", website="https://black.co"),
    ])
    p = _write_latin1(tmp_path, "Oregon Directory test.csv", csv_text)
    rec = OrCobidAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_city"] == "Portland"
    assert rec["address_state"] == "OR"
    assert rec["owner_name"] == "Jordan Smith"
    assert rec["phone"] == "503-555-0100"
    assert rec["email"] == "j@black.co"
    assert rec["website"] == "https://black.co"


def test_or_dedups_same_firm_across_cert_types(tmp_path):
    csv_text = _make_dir_csv("COBID Certified Firms Directory", _OR_HEADERS, [
        _or_row(name="Black Co", first="Jordan", last="Smith", location="Portland, OR", cert="ESB"),
        _or_row(name="Black Co", first="Jordan", last="Smith", location="Portland, OR", cert="MBE"),
    ])
    p = _write_latin1(tmp_path, "Oregon Directory test.csv", csv_text)
    records = OrCobidAdapter(file_path=p).run()
    assert len(records) == 1


def test_or_confidence_is_confirmed_black():
    assert OrCobidAdapter.CONFIDENCE == "confirmed_black"


# ── nv_dbe adapter tests ──────────────────────────────────────────────────────

from adapters.nv_dbe import NvDbeAdapter

_NV_HEADERS = ["Company Name", "DBA Name", "Owner First", "Owner Last",
               "Physical Address", "City", "State", "Zip",
               "Mailing Address", "City", "State", "Zip", "Phone", "Fax",
               "Email", "Website", "Agency", "Certification Type", "Ethnicity",
               "Gender", "Certified", "Capability", "County"]


def _nv_row(name="Acme Black LLC", first="Jordan", last="Smith",
            paddr="100 Main St", city="Las Vegas", state="NV", zipcode="\t89115",
            phone="702-555-0100", email="a@b.com", website="https://acme.com",
            cert="DBE", ethnicity="BLACK AMERICAN", capability="Trucking"):
    return [name, "", first, last, paddr, city, state, zipcode,
            paddr, city, state, zipcode, phone, "", email, website, "NDOT",
            cert, ethnicity, "Male", "2024", capability, "Clark"]


def test_nv_filters_black_american_case_insensitive(tmp_path):
    csv_text = _make_dir_csv("NDOT DBE Vendor List", _NV_HEADERS, [
        _nv_row(name="Black Co", ethnicity="BLACK AMERICAN"),
        _nv_row(name="Cauc Co", ethnicity="CAUCASIAN"),
        _nv_row(name="Hisp Co", ethnicity="HISPANIC AMERICAN"),
    ])
    p = _write_latin1(tmp_path, "Nevada Directory test.csv", csv_text)
    records = NvDbeAdapter(file_path=p).run()
    assert [r["business_name"] for r in records] == ["Black Co"]


def test_nv_maps_physical_address_and_strips_zip_tab(tmp_path):
    csv_text = _make_dir_csv("NDOT DBE Vendor List", _NV_HEADERS, [
        _nv_row(name="Black Co", paddr="742 D Street", city="Elko", state="NV",
                zipcode="\t89801", phone="775-555-0100", email="b@c.co"),
    ])
    p = _write_latin1(tmp_path, "Nevada Directory test.csv", csv_text)
    rec = NvDbeAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_street"] == "742 D Street"
    assert rec["address_city"] == "Elko"
    assert rec["address_state"] == "NV"
    assert rec["address_zip"] == "89801"   # leading tab stripped
    assert rec["owner_name"] == "Jordan Smith"
    assert rec["phone"] == "775-555-0100"


def test_nv_confidence_is_confirmed_black():
    assert NvDbeAdapter.CONFIDENCE == "confirmed_black"


# ── shared helpers for the 2026-06-13 file-based adapters ─────────────────────

def _write_utf8(tmp_path, name, headers, rows):
    """Write a flat CSV (header on row 0) as utf-8-sig, like the AR/CA/NC exports."""
    import csv as _csv
    p = tmp_path / name
    with open(p, "w", encoding="utf-8-sig", newline="") as f:
        w = _csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)
    return p


def _write_xlsx(tmp_path, name, headers, rows, sheet="Sheet1"):
    """Write an .xlsx with header on row 1, like the FL/VA exports."""
    import openpyxl as _ox
    wb = _ox.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    p = tmp_path / name
    wb.save(p)
    return p


# ── ar_mwbe adapter tests ─────────────────────────────────────────────────────

from adapters.ar_mwbe import ArMwbeAdapter

_AR_HEADERS = ["CompanyName", "BusinessDescription", "Phone", "Street", "City",
               "StateCode", "Zip", "VendorCategory", "CertificationNumber",
               "NaicsCode", "AasisVendorNumber", "ContactFirstName",
               "ContactLastName", "ContactTitle", "ContactPhone", "ContactEmail",
               "County", "BusinessDesignation", "Website", "OAN"]


def _ar_row(name="Acme Black LLC", category="African American", street="100 Main St",
            city="Little Rock", first="Jordan", last="Smith", email="a@b.co"):
    return [name, "Consulting", "501-555-0100", street, city, "AR", "72201",
            category, "", "541611", "", first, last, "Owner", "501-555-0100",
            email, "Pulaski", "", "https://acme.co", ""]


def test_ar_filters_both_spellings_of_african_american(tmp_path):
    p = _write_utf8(tmp_path, "Arkansas test.csv", _AR_HEADERS, [
        _ar_row(name="Hyphen Co", category="African-American"),
        _ar_row(name="Space Co", category="African American"),
        _ar_row(name="Women Co", category="Women-Owned"),
        _ar_row(name="Hisp Co", category="Hispanic American"),
    ])
    names = sorted(r["business_name"] for r in ArMwbeAdapter(file_path=p).run())
    assert names == ["Hyphen Co", "Space Co"]


def test_ar_maps_fields_and_owner(tmp_path):
    p = _write_utf8(tmp_path, "Arkansas test.csv", _AR_HEADERS, [
        _ar_row(name="Black Co", street="742 D St", city="Conway",
                first="Alex", last="Jones", email="alex@black.co"),
    ])
    rec = ArMwbeAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_street"] == "742 D St"
    assert rec["address_city"] == "Conway"
    assert rec["address_state"] == "AR"
    assert rec["owner_name"] == "Alex Jones"
    assert rec["email"] == "alex@black.co"
    assert rec["certification"] == "MBE"


def test_ar_dedups_on_company_and_street(tmp_path):
    p = _write_utf8(tmp_path, "Arkansas test.csv", _AR_HEADERS, [
        _ar_row(name="Black Co", street="1 A St"),
        _ar_row(name="Black Co", street="1 A St"),
    ])
    assert len(ArMwbeAdapter(file_path=p).run()) == 1


def test_ar_confidence_is_confirmed_black():
    assert ArMwbeAdapter.CONFIDENCE == "confirmed_black"


# ── nc_hub adapter tests ──────────────────────────────────────────────────────

from adapters.nc_hub import NcHubAdapter

_NC_HEADERS = ["Name", "MainContactName", "MainContactEmail", "MainContactPhone",
               "AddressLine1", "City", "State", "ZipCode", "County", "URL",
               "HUB", "HUBCategory"]


def _nc_row(name="Acme Black LLC", hub="Certified", category="Black",
            street="100 Main St", city="Raleigh", contact="Jordan Smith"):
    return [name, contact, "a@b.co", "919-555-0100", street, city, "NC",
            "27601", "Wake", "https://acme.co", hub, category]


def test_nc_filters_certified_black_only(tmp_path):
    p = _write_utf8(tmp_path, "North Carolina Vendor Details test.csv", _NC_HEADERS, [
        _nc_row(name="Good Co", hub="Certified", category="Black"),
        _nc_row(name="Blank Hub Co", hub="", category="Black"),
        _nc_row(name="NotCert Co", hub="Not Certified", category="Black"),
        _nc_row(name="Hisp Co", hub="Certified", category="Hispanic"),
    ])
    names = [r["business_name"] for r in NcHubAdapter(file_path=p).run()]
    assert names == ["Good Co"]


def test_nc_maps_fields(tmp_path):
    p = _write_utf8(tmp_path, "North Carolina Vendor Details test.csv", _NC_HEADERS, [
        _nc_row(name="Black Co", street="9 Oak Ave", city="Durham", contact="Pat Lee"),
    ])
    rec = NcHubAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["owner_name"] == "Pat Lee"
    assert rec["address_street"] == "9 Oak Ave"
    assert rec["address_city"] == "Durham"
    assert rec["address_state"] == "NC"
    assert rec["certification"] == "HUB"


def test_nc_dedups(tmp_path):
    p = _write_utf8(tmp_path, "North Carolina Vendor Details test.csv", _NC_HEADERS, [
        _nc_row(name="Black Co", street="1 A St"),
        _nc_row(name="Black Co", street="1 A St"),
    ])
    assert len(NcHubAdapter(file_path=p).run()) == 1


def test_nc_confidence_is_confirmed_black():
    assert NcHubAdapter.CONFIDENCE == "confirmed_black"


# ── ca_mbe adapter tests ──────────────────────────────────────────────────────

from adapters.ca_mbe import CaMbeAdapter

_CA_HEADERS = ["Vendor Name ", "Contact", "Contact Email", "Primary Address",
               "Active Certifications", "Ethnicity", "Gender", "Industries",
               "Account Email", "Market Area", "Contact Phone", "Business Activity",
               "Website"]


def _ca_row(name="Acme Black LLC", ethnicity="Black American",
            address="1730 N Wilton Place\n Los Angeles 90028\n CA Los Angeles\n",
            contact="Jordan Smith", certs="MBE\nWBE"):
    return [name, contact, "a@b.co", address, certs, ethnicity, "Female",
            "Consulting", "acct@b.co", "Local", "(310) 555-0100", "Goods", "https://acme.co"]


def test_ca_filters_black_american(tmp_path):
    p = _write_utf8(tmp_path, "Black_MBE test.csv", _CA_HEADERS, [
        _ca_row(name="Black Co", ethnicity="Black American"),
        _ca_row(name="Hisp Co", ethnicity="Hispanic American"),
    ])
    names = [r["business_name"] for r in CaMbeAdapter(file_path=p).run()]
    assert names == ["Black Co"]


def test_ca_parses_multiline_address(tmp_path):
    p = _write_utf8(tmp_path, "Black_MBE test.csv", _CA_HEADERS, [
        _ca_row(name="Black Co",
                address="1730 N Wilton Place\n Los Angeles 90028\n CA Los Angeles\n"),
    ])
    rec = CaMbeAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_street"] == "1730 N Wilton Place"
    assert rec["address_city"] == "Los Angeles"
    assert rec["address_zip"] == "90028"
    assert rec["address_state"] == "CA"
    assert rec["owner_name"] == "Jordan Smith"
    assert "MBE" in rec["certification"]


def test_ca_dedups_on_vendor_name(tmp_path):
    p = _write_utf8(tmp_path, "Black_MBE test.csv", _CA_HEADERS, [
        _ca_row(name="Black Co"), _ca_row(name="Black Co"),
    ])
    assert len(CaMbeAdapter(file_path=p).run()) == 1


def test_ca_confidence_is_confirmed_black():
    assert CaMbeAdapter.CONFIDENCE == "confirmed_black"


# ── fl_mbe adapter tests ──────────────────────────────────────────────────────

from adapters.fl_mbe import FlMbeAdapter

_FL_HEADERS = ["Vendor Name", "Contact", "Email", "Address", "City", "State",
               "Phone Number"]


def _fl_row(name="Acme Black LLC", contact="Jordan Smith", address="100 Main St",
            city="Miami"):
    return [name, contact, "a@b.co", address, city, "FL", "(305) 555-0100"]


def test_fl_combines_files_and_keeps_all_rows(tmp_path):
    p1 = _write_xlsx(tmp_path, "Florida A-G.xlsx", _FL_HEADERS,
                     [_fl_row(name="Alpha Co"), _fl_row(name="Bravo Co")], sheet="Vendors")
    p2 = _write_xlsx(tmp_path, "Florida H-M.xlsx", _FL_HEADERS,
                     [_fl_row(name="Mike Co")], sheet="Vendors")
    names = sorted(r["business_name"] for r in FlMbeAdapter(file_paths=[p1, p2]).run())
    assert names == ["Alpha Co", "Bravo Co", "Mike Co"]


def test_fl_maps_fields(tmp_path):
    p = _write_xlsx(tmp_path, "Florida A-G.xlsx", _FL_HEADERS,
                    [_fl_row(name="Black Co", contact="Pat Lee",
                             address="9 Palm Ave", city="Tampa")], sheet="Vendors")
    rec = FlMbeAdapter(file_paths=[p]).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["owner_name"] == "Pat Lee"
    assert rec["address_street"] == "9 Palm Ave"
    assert rec["address_city"] == "Tampa"
    assert rec["address_state"] == "FL"
    assert rec["certification"] == "MBE"


def test_fl_dedups_across_files(tmp_path):
    p1 = _write_xlsx(tmp_path, "Florida A-G.xlsx", _FL_HEADERS,
                     [_fl_row(name="Black Co", address="1 A St")], sheet="Vendors")
    p2 = _write_xlsx(tmp_path, "Florida H-M.xlsx", _FL_HEADERS,
                     [_fl_row(name="Black Co", address="1 A St")], sheet="Vendors")
    assert len(FlMbeAdapter(file_paths=[p1, p2]).run()) == 1


def test_fl_confidence_is_confirmed_black():
    assert FlMbeAdapter.CONFIDENCE == "confirmed_black"


# ── va_swam adapter tests ─────────────────────────────────────────────────────

from adapters.va_swam import VaSwamAdapter

# Two repeated blocks to exercise first-occurrence-wins (SWaM then MWAA).
_VA_HEADERS = ["Certification Type", "Business website", "Company Name",
               "Contact Name", "Contact Phone", "Contact Email", "Mailing Address",
               "Mailing City", "Mailing State", "Mailing Zip", "Business Category",
               "Ethnicity",
               # second (MWAA) block — duplicate names, ignored by first-wins
               "Company Name", "Mailing Zip", "Ethnicity"]


def _va_row(name="Acme Black LLC", ethnicity="Black or African American",
            zipc="23220", cert="Minority Owned", contact="Jordan Smith",
            blk2_name="", blk2_zip="", blk2_eth=""):
    return [cert, "https://acme.co", name, contact, "(804) 555-0100", "a@b.co",
            "100 Main St", "Richmond", "VA", zipc, "Consulting", ethnicity,
            blk2_name, blk2_zip, blk2_eth]


def test_va_filters_black_or_african_american(tmp_path):
    p = _write_xlsx(tmp_path, "Virginia Directory Listing Export-test.xlsx", _VA_HEADERS, [
        _va_row(name="Black Co", ethnicity="Black or African American"),
        _va_row(name="Asian Co", ethnicity="Asian American"),
        _va_row(name="Blank Co", ethnicity=""),
    ], sheet="Directory")
    names = [r["business_name"] for r in VaSwamAdapter(file_path=p).run()]
    assert names == ["Black Co"]


def test_va_reads_first_block_only(tmp_path):
    # First block has the real values; second block is junk and must be ignored.
    p = _write_xlsx(tmp_path, "Virginia Directory Listing Export-test.xlsx", _VA_HEADERS, [
        _va_row(name="Black Co", zipc="23220", contact="Pat Lee",
                blk2_name="WRONG", blk2_zip="00000", blk2_eth="WRONG"),
    ], sheet="Directory")
    rec = VaSwamAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_zip"] == "23220"
    assert rec["owner_name"] == "Pat Lee"
    assert rec["address_city"] == "Richmond"
    assert rec["certification"] == "Minority Owned"


def test_va_dedups_on_name_and_zip(tmp_path):
    p = _write_xlsx(tmp_path, "Virginia Directory Listing Export-test.xlsx", _VA_HEADERS, [
        _va_row(name="Black Co", zipc="23220"),
        _va_row(name="Black Co", zipc="23220"),
    ], sheet="Directory")
    assert len(VaSwamAdapter(file_path=p).run()) == 1


def test_va_confidence_is_confirmed_black():
    assert VaSwamAdapter.CONFIDENCE == "confirmed_black"


# ── B2Gnow family adapters (houston_obo, pa_ucp_dbe, atlanta_aabe) ────────────

from adapters.houston_obo import HoustonOboAdapter
from adapters.pa_ucp_dbe import PaUcpDbeAdapter
from adapters.atlanta_aabe import AtlantaAabeAdapter

# gob2g CSV header with the duplicate physical/mailing City/State/Zip columns.
_B2_HEADERS = ["Company Name", "DBA Name", "Owner First", "Owner Last",
               "Physical Address", "City", "State", "Zip",
               "Mailing Address", "City", "State", "Zip", "Phone", "Fax",
               "Email", "Website", "Agency", "Certification Type", "Ethnicity",
               "Gender", "Capability", "Category", "Commodity Codes"]


def _b2_row(name="Acme Black LLC", first="Jordan", last="Smith", paddr="100 Main St",
            city="Houston", state="TX", zipc="\t77033", phone="713-555-0100",
            email="a@b.co", cert="MBE", ethnicity="Black", cap="Consulting"):
    return [name, "", first, last, paddr, city, state, zipc,
            paddr, city, state, zipc, phone, "", email, "https://acme.co",
            "City", cert, ethnicity, "Male", cap, "Services", "541611"]


def test_b2gnow_csv_filters_black_and_black_american(tmp_path):
    text = _make_dir_csv("Certified Directory", _B2_HEADERS, [
        _b2_row(name="Black Co", ethnicity="Black"),
        _b2_row(name="Black Am Co", ethnicity="Black American"),
        _b2_row(name="Cauc Co", ethnicity="Caucasian"),
        _b2_row(name="Asian Co", ethnicity="Asian"),
    ])
    p = _write_latin1(tmp_path, "Houston test Directory.csv", text)
    names = sorted(r["business_name"] for r in HoustonOboAdapter(file_path=p).run())
    assert names == ["Black Am Co", "Black Co"]


def test_b2gnow_csv_maps_fields_strips_zip_tab_and_owner(tmp_path):
    text = _make_dir_csv("Certified Directory", _B2_HEADERS, [
        _b2_row(name="Black Co", first="Pat", last="Lee", paddr="9 Oak Ave",
                city="Houston", state="TX", zipc="\t77002", email="pat@black.co"),
    ])
    p = _write_latin1(tmp_path, "Houston test Directory.csv", text)
    rec = HoustonOboAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_street"] == "9 Oak Ave"
    assert rec["address_city"] == "Houston"
    assert rec["address_state"] == "TX"
    assert rec["address_zip"] == "77002"      # leading tab stripped by map_record
    assert rec["owner_name"] == "Pat Lee"
    assert rec["email"] == "pat@black.co"


def test_b2gnow_csv_dedups_on_company_and_physical_address(tmp_path):
    text = _make_dir_csv("Certified Directory", _B2_HEADERS, [
        _b2_row(name="Black Co", paddr="1 A St", ethnicity="Black"),
        _b2_row(name="Black Co", paddr="1 A St", ethnicity="Black"),  # same firm, another commodity
    ])
    p = _write_latin1(tmp_path, "Houston test Directory.csv", text)
    assert len(HoustonOboAdapter(file_path=p).run()) == 1


def test_pa_ucp_filters_black_american(tmp_path):
    text = _make_dir_csv("Certified Directory", _B2_HEADERS, [
        _b2_row(name="PA Black Co", state="PA", zipc="\t19103", ethnicity="Black American"),
        _b2_row(name="PA Cauc Co", state="PA", zipc="\t19103", ethnicity="Caucasian"),
    ])
    p = _write_latin1(tmp_path, "Pennsylvania test Directory.csv", text)
    recs = PaUcpDbeAdapter(file_path=p).run()
    assert [r["business_name"] for r in recs] == ["PA Black Co"]
    assert recs[0]["certification"]  # cert type carried through


def test_b2gnow_confidence_is_confirmed_black():
    assert HoustonOboAdapter.CONFIDENCE == "confirmed_black"
    assert PaUcpDbeAdapter.CONFIDENCE == "confirmed_black"
    assert AtlantaAabeAdapter.CONFIDENCE == "confirmed_black"


# Atlanta — B2Gnow ".xls" that is actually an HTML <table>.
_ATL_HEADERS = ["Company Name", "DBA Name", "Owner First", "Owner Last", "Location",
                "Phone", "Fax", "Email", "Website", "Agency", "Certification Type",
                "Expiration", "Capability", "Market Area", "Supplier ID#",
                "Commodity Codes"]


def _make_b2gnow_html(headers, rows):
    def tr(cells):
        return "<tr>" + "".join(f"<td>{c}</td>" for c in cells) + "</tr>\n"
    body = "<strong>Certified Directory</strong><br/>\nAs of 6/14/2026<br/>\n"
    body += "<table border='1'>\n" + tr(headers)
    for r in rows:
        body += tr(r)
    body += "</table>"
    return body


def _atl_row(name="Acme Black LLC", first="Jordan", last="Smith",
             location="Atlanta, GA", phone="404-555-0100", email="a@b.co",
             cert="AABE", cap="Electrical"):
    return [name, "", first, last, location, phone, "404-555-0101", email,
            "https://acme.co", "Atlanta", cert, "3/11/2029", cap, "", "1303394",
            "238210 - Electrical"]


def test_atlanta_filters_aabe_only(tmp_path):
    html = _make_b2gnow_html(_ATL_HEADERS, [
        _atl_row(name="AABE Co", cert="AABE"),
        _atl_row(name="SBE Co", cert="SBE"),
        _atl_row(name="FBE Co", cert="FBE"),
    ])
    p = _write_latin1(tmp_path, "Atlanta test Directory.xls", html)
    names = [r["business_name"] for r in AtlantaAabeAdapter(file_path=p).run()]
    assert names == ["AABE Co"]


def test_atlanta_parses_location_and_fields(tmp_path):
    html = _make_b2gnow_html(_ATL_HEADERS, [
        _atl_row(name="Black Co", first="Pat", last="Lee", location="Decatur, GA",
                 email="pat@black.co", cert="AABE"),
    ])
    p = _write_latin1(tmp_path, "Atlanta test Directory.xls", html)
    rec = AtlantaAabeAdapter(file_path=p).run()[0]
    assert rec["business_name"] == "Black Co"
    assert rec["address_city"] == "Decatur"
    assert rec["address_state"] == "GA"
    assert rec["owner_name"] == "Pat Lee"
    assert rec["email"] == "pat@black.co"
    assert rec["certification"] == "AABE"


def test_atlanta_dedups_on_company_and_location(tmp_path):
    html = _make_b2gnow_html(_ATL_HEADERS, [
        _atl_row(name="Black Co", location="Atlanta, GA"),
        _atl_row(name="Black Co", location="Atlanta, GA"),
    ])
    p = _write_latin1(tmp_path, "Atlanta test Directory.xls", html)
    assert len(AtlantaAabeAdapter(file_path=p).run()) == 1
